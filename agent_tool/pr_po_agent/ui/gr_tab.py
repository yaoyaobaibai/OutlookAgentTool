# -*- coding: utf-8 -*-
"""PR/PO Agent - GR-Acubuy Tab UI skeleton (v1.3.2+).

Functional stub: form fields, attachments placeholder, Save Draft button.
All backend integration is deferred to Phase 2 (per .omo/plans/gr-acubuy-team-plan.md).
"""

import logging
import os
import re
import shutil
import subprocess
import sys
import time
import tkinter as tk
from pathlib import Path
from tkinter import ttk, messagebox, filedialog

from config import (
    ATTACHMENT_CATEGORIES,
    DEFAULT_FONT,
    DEFAULT_FONT_BOLD,
    EXAMPLE_TASKS,
    EXCEL_ATTACH_COLUMNS,
    GR_ACUBUY_UI_TEXT,
    PDF_MERGE_TOOL_PATH,
    PRIORITY_DISPLAY,
    STATS,
    STATS_LABELS,
    STATUS_DISPLAY,
)

logger = logging.getLogger(__name__)


class GrAcubuyTab:
    """GR-Acubuy main window tab — functional stub (Phase 1 UI skeleton)."""

    def __init__(self, parent):
        """Build the GR-Acubuy tab inside the parent notebook."""
        self.parent = parent
        self.frame = ttk.Frame(parent, padding=(10, 5))

        # 2-row grid: row 0 = scrollable body (weight=1), row 1 = pinned footer (weight=0)
        self.frame.grid_rowconfigure(0, weight=1)
        self.frame.grid_rowconfigure(1, weight=0)
        self.frame.grid_columnconfigure(0, weight=1)

        # Field attr init (set by builders)
        self._purchase_order_entry = None
        self._delivery_note_entry = None
        self._internal_comment_text = None
        self._quantity_received_spin = None
        self._requestor_entry = None
        self._approver_2_entry = None
        self._attachments_list = None
        self._status_label = None
        self._generate_btn = None
        self._merge_pdfs_btn = None

        # Attachment state: list of dicts {category, src_path, description, is_pdf, dest_relative_path}
        self._attachments = []

        # --- Row 0: scrollable body (Canvas + Scrollbar + embedded Frame) ---
        self._body_canvas = tk.Canvas(self.frame, borderwidth=0, highlightthickness=0)
        self._body_scrollbar = ttk.Scrollbar(self.frame, orient="vertical",
                                             command=self._body_canvas.yview)
        self._scroll_body = ttk.Frame(self._body_canvas)

        self._body_canvas_window = self._body_canvas.create_window(
            (0, 0), window=self._scroll_body, anchor="nw")

        self._body_canvas.configure(yscrollcommand=self._body_scrollbar.set)

        self._body_canvas.grid(row=0, column=0, sticky="nsew")
        self._body_scrollbar.grid(row=0, column=1, sticky="ns")

        # Bind configure events for scrollregion + width sync
        self._body_canvas.bind("<Configure>", self._on_canvas_configure)
        self._scroll_body.bind("<Configure>", self._on_scroll_body_configure)

        # --- Row 1: pinned footer ---
        self._footer = ttk.Frame(self.frame)
        self._footer.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Build sections into scroll body
        self._build_today_overview()
        self._build_form_section()
        self._build_attachments_section()

        # Build sections into footer
        self._build_action_section(parent=self._footer)
        self._build_status_row(parent=self._footer)

        # Windows mouse-wheel bindings (active only while pointer is over body)
        self._scroll_body.bind("<Enter>", self._on_body_enter)
        self._scroll_body.bind("<Leave>", self._on_body_leave)
        self.frame.bind("<Destroy>", self._on_destroy)

    # ------------------------------------------------------------------
    # 今日概览 section (stats + task list, was in main_window)
    # ------------------------------------------------------------------

    def _build_today_overview(self):
        overview = ttk.LabelFrame(
            self._scroll_body,
            text=GR_ACUBUY_UI_TEXT["today_overview_title"],
            padding=(10, 5),
        )
        overview.pack(fill="x", pady=(0, 8))

        # 3 stat cards (smaller than original main_window)
        cards_frame = ttk.Frame(overview)
        cards_frame.pack(fill="x", pady=(0, 5))
        for idx, (key, value) in enumerate(STATS.items()):
            card = ttk.LabelFrame(
                cards_frame,
                text=STATS_LABELS[key],
                padding=(8, 4),
            )
            card.grid(row=0, column=idx, padx=(0, 6), sticky="nsew")
            num_label = ttk.Label(card, text=str(value), font=("Microsoft YaHei", 14, "bold"))
            num_label.pack(fill="x")
            cards_frame.columnconfigure(idx, weight=1, uniform="stat")

        # Compact task list (reuse EXAMPLE_TASKS as placeholder)
        list_frame = ttk.Frame(overview)
        list_frame.pack(fill="x")
        columns = ("id", "title", "status")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=4)
        tree.heading("id", text="ID")
        tree.heading("title", text="\u6807\u9898")
        tree.heading("status", text="\u72b6\u6001")
        tree.column("id", width=60, anchor="center")
        tree.column("title", width=380)
        tree.column("status", width=80, anchor="center")
        for task in EXAMPLE_TASKS:
            tree.insert("", "end", values=(
                task["id"], task["title"], STATUS_DISPLAY[task["status"]],
            ))
        tree.pack(fill="x")

    # ------------------------------------------------------------------
    # GR 表单 section (6 fields mapped to gr_template.xlsx columns)
    # ------------------------------------------------------------------

    def _build_form_section(self):
        """Build the GR Acubuy form section with 6 Excel-mapped fields.

        User fills these, clicks Generate Excel, script writes the values
        to inbox/<GR_<timestamp>.xlsx (from gr_template.xlsx). Colleague C's
        acubuy_agent reads that Excel.
        """
        form = ttk.LabelFrame(
            self._scroll_body,
            text=GR_ACUBUY_UI_TEXT["form_section_title"],
            padding=(10, 8),
        )
        form.pack(fill="x", pady=(0, 8))

        # Purchase Order
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_purchase_order_label"], anchor="w", width=28).pack(side="left", padx=(0, 6))
        self._purchase_order_entry = ttk.Entry(row)
        self._purchase_order_entry.pack(side="left", fill="x", expand=True)

        # Delivery Note
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_delivery_note_label"], anchor="w", width=28).pack(side="left", padx=(0, 6))
        self._delivery_note_entry = ttk.Entry(row)
        self._delivery_note_entry.pack(side="left", fill="x", expand=True)

        # Internal Comment (multiline)
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_internal_comment_label"], anchor="w", width=28).pack(side="left", padx=(0, 6), anchor="nw")
        self._internal_comment_text = tk.Text(row, height=3, width=40)
        self._internal_comment_text.pack(side="left", fill="x", expand=True)

        # Quantity Received (integer)
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_quantity_received_label"], anchor="w", width=28).pack(side="left", padx=(0, 6))
        self._quantity_received_spin = ttk.Spinbox(row, from_=0, to=99999, width=10)
        self._quantity_received_spin.pack(side="left", fill="x", expand=True)

        # Requestor
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_requestor_label"], anchor="w", width=28).pack(side="left", padx=(0, 6))
        self._requestor_entry = ttk.Entry(row)
        self._requestor_entry.pack(side="left", fill="x", expand=True)

        # Approver 2 (Min Band E)
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=GR_ACUBUY_UI_TEXT["gr_approver_2_label"], anchor="w", width=28).pack(side="left", padx=(0, 6))
        self._approver_2_entry = ttk.Entry(row)
        self._approver_2_entry.pack(side="left", fill="x", expand=True)

    # ------------------------------------------------------------------
    # 附件 section
    # ------------------------------------------------------------------

    def _build_attachments_section(self):
        att = ttk.LabelFrame(
            self._scroll_body,
            text=GR_ACUBUY_UI_TEXT["attachments_section_title"],
            padding=(10, 5),
        )
        att.pack(fill="x", pady=(0, 8))

        list_frame = ttk.Frame(att)
        list_frame.pack(fill="x")
        self._attachments_list = tk.Listbox(list_frame, height=4, selectmode="extended")
        self._attachments_list.pack(side="left", fill="x", expand=True)
        self._attachments_list.insert("end", GR_ACUBUY_UI_TEXT["gr_no_attachments"])
        self._attachments_list.bind("<Double-1>", self._on_remove_attachment)

        # Button frame on the right side
        btn_frame = ttk.Frame(list_frame)
        btn_frame.pack(side="right", padx=(6, 0))

        self._merge_pdfs_btn = ttk.Button(
            btn_frame,
            text=GR_ACUBUY_UI_TEXT["gr_merge_pdfs_btn"],
            command=self._on_merge_pdfs,
            state="disabled",
            width=12,
        )
        self._merge_pdfs_btn.pack(side="top", pady=(0, 4))

        remove_btn = ttk.Button(
            btn_frame,
            text=GR_ACUBUY_UI_TEXT["gr_remove_attachment_btn"],
            command=self._on_remove_attachment,
            width=12,
        )
        remove_btn.pack(side="top", pady=(0, 4))

        add_btn = ttk.Button(
            btn_frame,
            text=GR_ACUBUY_UI_TEXT["gr_add_attachment_btn"],
            command=self._on_add_attachment,
            width=12,
        )
        add_btn.pack(side="top")

    # ------------------------------------------------------------------
    # 操作 section (Save Draft button)
    # ------------------------------------------------------------------

    def _build_action_section(self, parent=None):
        if parent is None:
            parent = self._scroll_body
        action = ttk.LabelFrame(
            parent,
            text=GR_ACUBUY_UI_TEXT["action_section_title"],
            padding=(10, 5),
        )
        action.pack(fill="x", pady=(0, 8))
        btn_row = ttk.Frame(action)
        btn_row.pack(anchor="center", pady=4)
        self._generate_btn = ttk.Button(
            btn_row,
            text=GR_ACUBUY_UI_TEXT["gr_generate_excel_btn"],
            command=self._on_generate_excel,
            width=20,
        )
        self._generate_btn.pack(side="left", padx=10)
        self._auto_fetch_btn = ttk.Button(
            btn_row,
            text=GR_ACUBUY_UI_TEXT["gr_auto_fetch_btn"],
            command=self._on_auto_fetch,
            width=20,
        )
        self._auto_fetch_btn.pack(side="left", padx=10)

    # ------------------------------------------------------------------
    # 状态 row
    # ------------------------------------------------------------------

    def _build_status_row(self, parent=None):
        if parent is None:
            parent = self._scroll_body
        status = ttk.Frame(parent, padding=(10, 3))
        status.pack(fill="x")
        self._status_label = ttk.Label(
            status,
            text=GR_ACUBUY_UI_TEXT["gr_status_disconnected"],
            foreground="gray",
        )
        self._status_label.pack(side="left")

    # ------------------------------------------------------------------
    # Status helper
    # ------------------------------------------------------------------

    def _set_status(self, text, color):
        """Set persistent status label text and foreground color."""
        self._status_label.config(text=text, foreground=color)

    def _resolve_template_path(self):
        """Resolve gr_template.xlsx in fixed priority order.

        Returns the first existing Path or None if no candidate exists.
        1. frozen: _MEIPASS/docs/templates/gr_template.xlsx
        2. source: workspace root docs/templates/gr_template.xlsx
        3. fallback: ~/PRPOAgent_docs/templates/gr_template.xlsx
        """
        frozen = getattr(sys, "frozen", False)
        meipass = getattr(sys, "_MEIPASS", None)

        if frozen and meipass:
            candidate = Path(meipass) / "docs" / "templates" / "gr_template.xlsx"
            if candidate.is_file():
                return candidate

        candidate = Path(__file__).resolve().parents[3] / "docs" / "templates" / "gr_template.xlsx"
        if candidate.is_file():
            return candidate

        candidate = Path(os.path.expandvars(os.path.expanduser("~"))) / "PRPOAgent_docs" / "templates" / "gr_template.xlsx"
        if candidate.is_file():
            return candidate

        return None

    # ------------------------------------------------------------------
    # Canvas / scroll lifecycle
    # ------------------------------------------------------------------

    def _on_canvas_configure(self, event):
        """Sync embedded frame width to canvas viewport on resize."""
        if event.width > 1:
            self._body_canvas.itemconfig(self._body_canvas_window, width=event.width)

    def _on_scroll_body_configure(self, event):
        """Update scrollregion and sync width when content changes."""
        self._body_canvas.configure(scrollregion=self._body_canvas.bbox("all"))
        canvas_width = self._body_canvas.winfo_width()
        if canvas_width > 1:
            self._body_canvas.itemconfig(self._body_canvas_window, width=canvas_width)

    def _on_body_enter(self, event):
        """Bind mousewheel to canvas scrolling when pointer enters body."""
        self._root_window = self.frame.winfo_toplevel()
        self._root_window.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_body_leave(self, event):
        """Unbind mousewheel when pointer leaves body."""
        if hasattr(self, "_root_window"):
            self._root_window.unbind_all("<MouseWheel>")
            del self._root_window

    def _on_mousewheel(self, event):
        """Scroll canvas on Windows mouse wheel (guard: no-op when content fits)."""
        bbox = self._body_canvas.bbox("all")
        if bbox and bbox[3] > self._body_canvas.winfo_height():
            self._body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_destroy(self, event):
        """Clean up wheel bindings on destroy."""
        try:
            self.frame.winfo_toplevel().unbind_all("<MouseWheel>")
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Attachment handlers (v1.3.3+)
    # ------------------------------------------------------------------

    def _on_add_attachment(self, event=None):
        """Open file dialog, then category/description dialog for each file."""
        paths = filedialog.askopenfilenames(
            title="Select Attachments",
            filetypes=[("All files", "*.*")],
            initialdir=os.path.expanduser("~"),
        )
        if not paths:
            return

        for path in paths:
            result = self._attachment_dialog(path)
            if result is None:
                continue  # user cancelled this file
            category, description = result
            is_pdf = path.lower().endswith(".pdf")
            self._attachments.append({
                "category": category,
                "src_path": path,
                "description": description,
                "is_pdf": is_pdf,
                "dest_relative_path": None,
            })

        self._refresh_attachments_listbox()
        self._refresh_merge_btn_state()

    def _attachment_dialog(self, path):
        """Show Toplevel dialog for category + description. Returns (category, description) or None."""
        dialog = tk.Toplevel(self.frame)
        dialog.title(GR_ACUBUY_UI_TEXT["gr_attachment_dialog_title"])
        dialog.transient(self.frame)
        dialog.grab_set()
        dialog.resizable(False, False)

        # File label
        file_frame = ttk.Frame(dialog, padding=(10, 10, 10, 0))
        file_frame.pack(fill="x")
        ttk.Label(file_frame, text=GR_ACUBUY_UI_TEXT["gr_attachment_file_label"]).pack(side="left")
        ttk.Label(file_frame, text=os.path.basename(path), foreground="gray").pack(side="left", padx=(6, 0))

        # Category
        cat_frame = ttk.Frame(dialog, padding=(10, 6, 10, 0))
        cat_frame.pack(fill="x")
        ttk.Label(cat_frame, text=GR_ACUBUY_UI_TEXT["gr_attachment_category_label"]).pack(side="left")
        cat_var = tk.StringVar(value=ATTACHMENT_CATEGORIES[-1])  # default "Other"
        cat_combo = ttk.Combobox(
            cat_frame, textvariable=cat_var,
            values=ATTACHMENT_CATEGORIES, state="readonly", width=22,
        )
        cat_combo.pack(side="left", padx=(6, 0))

        # Description
        desc_frame = ttk.Frame(dialog, padding=(10, 6, 10, 0))
        desc_frame.pack(fill="x")
        ttk.Label(desc_frame, text=GR_ACUBUY_UI_TEXT["gr_attachment_desc_label"]).pack(side="left", anchor="nw")
        desc_text = tk.Text(desc_frame, height=3, width=40)
        desc_text.pack(side="left", fill="x", expand=True, padx=(6, 0))

        result_holder = []

        def on_ok():
            result_holder.append((cat_var.get(), desc_text.get("1.0", tk.END).strip()))
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        btn_frame = ttk.Frame(dialog, padding=(10, 10))
        btn_frame.pack(fill="x")
        ttk.Button(btn_frame, text="OK", command=on_ok, width=10).pack(side="right", padx=(4, 0))
        ttk.Button(btn_frame, text="Cancel", command=on_cancel, width=10).pack(side="right")

        dialog.wait_window()
        return result_holder[0] if result_holder else None

    def _on_remove_attachment(self, event=None):
        """Remove selected attachments from list."""
        selection = self._attachments_list.curselection()
        if not selection:
            return
        # Sort descending so pop indices don't shift
        for idx in sorted(selection, reverse=True):
            try:
                self._attachments.pop(idx)
            except IndexError:
                pass
        self._refresh_attachments_listbox()
        self._refresh_merge_btn_state()

    def _refresh_attachments_listbox(self):
        """Rebuild Listbox from self._attachments."""
        self._attachments_list.delete(0, "end")
        if not self._attachments:
            self._attachments_list.insert("end", GR_ACUBUY_UI_TEXT["gr_no_attachments"])
            return
        for item in self._attachments:
            display = f"{os.path.basename(item['src_path'])} [{item['category']}]"
            self._attachments_list.insert("end", display)

    def _refresh_merge_btn_state(self):
        """Enable merge button only when >= 2 PDFs."""
        pdf_count = sum(1 for a in self._attachments if a["is_pdf"])
        if self._merge_pdfs_btn is not None:
            state = "normal" if pdf_count >= 2 else "disabled"
            self._merge_pdfs_btn.configure(state=state)

    def _on_merge_pdfs(self):
        """Merge selected PDFs via PDFMergeTool.exe subprocess."""
        pdf_paths = [a["src_path"] for a in self._attachments if a["is_pdf"]]
        if len(pdf_paths) < 2:
            messagebox.showwarning(
                "Acubuy", GR_ACUBUY_UI_TEXT["gr_no_pdf_selected_for_merge"],
            )
            return

        # Ask for output path
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        out_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=f"merged_{timestamp}.pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Save Merged PDF As",
        )
        if not out_path:
            return

        # Locate PDFMergeTool.exe
        pdf_merge_tool = None
        candidate1 = os.path.expanduser("~/PDFMergeTool.exe")
        if os.path.isfile(candidate1):
            pdf_merge_tool = candidate1
        elif getattr(sys, "frozen", False):
            candidate2 = os.path.join(os.path.dirname(sys.executable), "PDFMergeTool.exe")
            if os.path.isfile(candidate2):
                pdf_merge_tool = candidate2

        if pdf_merge_tool is None:
            messagebox.showerror(
                "Acubuy",
                "PDFMergeTool.exe not found.\n\n"
                "Expected locations:\n"
                f"  1. {candidate1}\n"
                f"  2. (frozen) {os.path.join(os.path.dirname(sys.executable), 'PDFMergeTool.exe')}\n\n"
                "Place PDFMergeTool.exe in one of these locations.",
            )
            return

        try:
            proc = subprocess.run(
                [pdf_merge_tool, "--auto", "--output", out_path, "--files"] + pdf_paths,
                capture_output=True,
                creationflags=0x08000000,
            )

            def _safe_decode(data):
                """Decode bytes: try utf-8 -> gbk -> latin-1."""
                for enc in ("utf-8", "gbk", "latin-1"):
                    try:
                        return data.decode(enc)
                    except (UnicodeDecodeError, LookupError):
                        continue
                return data.decode("latin-1", errors="replace")

            if proc.returncode == 0:
                merged_attachment = {
                    "category": "Other",
                    "src_path": out_path,
                    "description": f"Merged from {len(pdf_paths)} PDFs",
                    "is_pdf": True,
                    "dest_relative_path": None,
                }
                self._attachments.append(merged_attachment)
                self._refresh_attachments_listbox()
                self._refresh_merge_btn_state()
                messagebox.showinfo("Acubuy", f"PDFs merged successfully:\n{out_path}")
                logger.info("Merged %d PDFs -> %s", len(pdf_paths), out_path)
            else:
                stderr_msg = _safe_decode(proc.stderr) if proc.stderr else "Unknown error"
                messagebox.showerror("Acubuy", f"PDF merge failed:\n{stderr_msg}")
                logger.error("PDF merge failed: %s", stderr_msg)
        except Exception as e:
            messagebox.showerror("Acubuy", f"PDF merge error: {e}")
            logger.error("PDF merge error: %s", e, exc_info=True)

    def _on_generate_excel(self):
        """Generate Excel from UI fields, save to inbox/ for colleague C to read.

        Flow:
        1. Collect UI field values
        2. Validate (check required fields)
        3. Load gr_template.xlsx (template)
        4. Write new row with UI values
        5. Save to inbox/GR_<timestamp>.xlsx
        6. Show success message with file path
        """
        # 0. Set in-progress status (gray) before anything else
        self._set_status(GR_ACUBUY_UI_TEXT["gr_status_in_progress"], "gray")

        # 1. Collect form values
        form_data = {
            GR_ACUBUY_UI_TEXT["gr_form_key_purchase_order"]:   self._purchase_order_entry.get().strip(),
            GR_ACUBUY_UI_TEXT["gr_form_key_delivery_note"]:    self._delivery_note_entry.get().strip(),
            GR_ACUBUY_UI_TEXT["gr_form_key_internal_comment"]: self._internal_comment_text.get("1.0", tk.END).strip(),
            GR_ACUBUY_UI_TEXT["gr_form_key_quantity_received"]: self._quantity_received_spin.get().strip(),
            GR_ACUBUY_UI_TEXT["gr_form_key_requestor"]:        self._requestor_entry.get().strip(),
            GR_ACUBUY_UI_TEXT["gr_form_key_approver_2"]:       self._approver_2_entry.get().strip(),
        }

        # 2. Validate required fields
        required = [
            GR_ACUBUY_UI_TEXT["gr_form_key_purchase_order"],
            GR_ACUBUY_UI_TEXT["gr_form_key_delivery_note"],
            GR_ACUBUY_UI_TEXT["gr_form_key_quantity_received"],
            GR_ACUBUY_UI_TEXT["gr_form_key_requestor"],
            GR_ACUBUY_UI_TEXT["gr_form_key_approver_2"],
        ]
        missing = [k for k in required if not form_data.get(k)]
        if missing:
            self._set_status(GR_ACUBUY_UI_TEXT["gr_failure_status"], "red")
            messagebox.showerror(
                "Acubuy",
                f"Required fields missing: {', '.join(missing)}\n\nPlease fill all required fields.",
            )
            logger.warning("Generate Excel failed: missing fields %s", missing)
            return

        # 3-5. Generate Excel
        try:
            from openpyxl import load_workbook

            template_path = self._resolve_template_path()
            if template_path is None:
                self._set_status(GR_ACUBUY_UI_TEXT["gr_failure_status"], "red")
                messagebox.showerror(
                    "Acubuy",
                    "Template not found in any expected location.\n\n"
                    "Checked:\n"
                    "1. frozen _MEIPASS/docs/templates/gr_template.xlsx\n"
                    "2. workspace docs/templates/gr_template.xlsx\n"
                    "3. ~/PRPOAgent_docs/templates/gr_template.xlsx\n"
                    f"Current __file__ = {__file__}\n"
                    "Reinstall or place template in the workspace docs/templates/ directory.",
                )
                return

            wb = load_workbook(str(template_path))
            ws = wb.active

            # Clear sample data rows (keep only the header at row 1)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            # Resolve inbox path and safe filename
            inbox_dir = os.path.expandvars(os.path.expanduser("~/PRPOAgent/inbox"))
            os.makedirs(inbox_dir, exist_ok=True)
            po_value = form_data[GR_ACUBUY_UI_TEXT["gr_form_key_purchase_order"]]
            safe_po = re.sub(r'[<>:"/\\|?*]', '_', po_value)
            output_path = os.path.join(inbox_dir, f"GR_{safe_po}.xlsx")

            # Copy attachments to inbox/attachments/<safe_PO>/
            attach_dest_dir = os.path.join(inbox_dir, "attachments", safe_po)
            os.makedirs(attach_dest_dir, exist_ok=True)
            missing_files = []
            prepared_attachments = []  # list of (category, rel_path, description)
            for att in self._attachments:
                src = att["src_path"]
                if not os.path.exists(src):
                    missing_files.append(src)
                    continue
                filename = os.path.basename(src)
                dest = os.path.join(attach_dest_dir, filename)
                counter = 1
                base, ext = os.path.splitext(filename)
                while os.path.exists(dest):
                    dest = os.path.join(attach_dest_dir, f"{base}_{counter}{ext}")
                    counter += 1
                try:
                    shutil.copy2(src, dest)
                    rel_path = os.path.relpath(dest, inbox_dir).replace(os.sep, "/")
                    prepared_attachments.append((att["category"], rel_path, att["description"]))
                except Exception as e:
                    logger.error("Copy attachment failed: %s -> %s: %s", src, dest, e)
                    missing_files.append(src)

            # Append data row
            ws.append([
                form_data[GR_ACUBUY_UI_TEXT["gr_form_key_purchase_order"]],
                form_data[GR_ACUBUY_UI_TEXT["gr_form_key_delivery_note"]],
                form_data[GR_ACUBUY_UI_TEXT["gr_form_key_internal_comment"]],
                int(form_data[GR_ACUBUY_UI_TEXT["gr_form_key_quantity_received"]] or 0),
                form_data[GR_ACUBUY_UI_TEXT["gr_form_key_requestor"]],
                form_data[GR_ACUBUY_UI_TEXT["gr_form_key_approver_2"]],
            ])

            # Append attachment rows (one per prepared attachment)
            for category, rel_path, description in prepared_attachments:
                ws.append([
                    form_data[GR_ACUBUY_UI_TEXT["gr_form_key_purchase_order"]],
                    form_data[GR_ACUBUY_UI_TEXT["gr_form_key_delivery_note"]],
                    "",  # Internal Comment
                    0,   # Quantity Received
                    "",  # Requestor
                    "",  # Approver 2
                    category,
                    rel_path,
                    description,
                ])

            # Save to inbox/
            wb.save(output_path)

            # Warn about missing/copy-failed attachments
            if missing_files:
                messagebox.showwarning(
                    "Acubuy",
                    GR_ACUBUY_UI_TEXT["gr_file_copy_failed_warning"].format(
                        paths="\n".join(missing_files),
                        po=safe_po,
                        dir=attach_dest_dir,
                    ),
                )

            # 6. Success feedback: green persistent status + messagebox
            self._set_status(
                GR_ACUBUY_UI_TEXT["gr_success_status_prefix"] + os.path.basename(output_path),
                "green",
            )
            messagebox.showinfo(
                GR_ACUBUY_UI_TEXT["gr_success_title"],
                GR_ACUBUY_UI_TEXT["gr_success_body"].format(path=output_path),
            )
            logger.info("Generated GR Excel: %s", output_path)
        except ImportError:
            self._set_status(GR_ACUBUY_UI_TEXT["gr_failure_status"], "red")
            messagebox.showerror(
                "Acubuy",
                "openpyxl not installed. Run: pip install openpyxl",
            )
            logger.error("openpyxl not installed")
        except Exception as e:
            self._set_status(GR_ACUBUY_UI_TEXT["gr_failure_status"], "red")
            messagebox.showerror("Acubuy", f"Generate Excel failed: {e}")
            logger.error("Generate Excel failed: %s", e, exc_info=True)

    def _on_auto_fetch(self):
        """Stub: future functionality to import GR form data from email/attachments."""
        messagebox.showinfo("Acubuy", GR_ACUBUY_UI_TEXT["gr_auto_fetch_stub_msg"])
