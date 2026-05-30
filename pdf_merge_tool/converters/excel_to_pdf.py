"""Excel 文件转 PDF"""
import os
import subprocess
import sys
from typing import Optional

_CREATE_NO_WINDOW = 0x08000000 if sys.platform == 'win32' else 0


def excel_to_pdf(
    excel_path: str,
    output_dir: str,
    page_size: tuple = None
) -> tuple[bool, str]:
    """
    将 Excel 文件转换为 PDF
    
    优先使用 MS Excel，如果不可用则尝试 LibreOffice
    
    Args:
        excel_path: Excel 文件路径
        output_dir: 输出目录
    
    Returns:
        (是否成功, 错误消息)
    """
    # 方案1: 使用 MS Excel (Windows + Office)
    if os.name == 'nt':
        success, error = _excel_to_pdf_com(excel_path, output_dir, page_size=page_size)
        if success:
            return True, ""
        # Log COM failure reason before trying LibreOffice
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Excel COM 转换失败（{os.path.basename(excel_path)}），尝试 LibreOffice: {error}")

    # 方案2: 使用 LibreOffice
    success, error = _excel_to_pdf_libreoffice(excel_path, output_dir, page_size=page_size)
    if success:
        return True, ""
    
    # 方案3: 使用 openpyxl + Pillow (最终备用方案)
    import logging
    logger = logging.getLogger(__name__)
    logger.warning(f"LibreOffice 转换失败（{os.path.basename(excel_path)}），尝试 openpyxl+Pillow 备用方案: {error}")
    
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_pdf = os.path.join(output_dir, f"{base_name}.pdf")
    success, error = excel_to_pdf_images(excel_path, output_pdf, page_size=page_size)
    if success:
        logger.info(f"Excel 通过 openpyxl+Pillow 转换成功: {os.path.basename(excel_path)}")
        return True, ""
    
    return False, error or "Excel 转 PDF 失败"


def _excel_to_pdf_com(
    excel_path: str,
    output_dir: str,
    page_size: tuple = None
) -> tuple[bool, str]:
    """
    使用 MS Excel COM 接口转换 Excel 到 PDF (Windows)
    
    Args:
        excel_path: Excel 文件路径
        output_dir: 输出目录
    
    Returns:
        (是否成功, 错误消息)
    """
    try:
        import win32com.client
        import pythoncom
        import logging
        logger = logging.getLogger(__name__)
        
        # 初始化 COM
        pythoncom.CoInitialize()
        
        try:
            # 获取输出 PDF 路径
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            output_pdf = os.path.join(output_dir, f"{base_name}.pdf")
            
            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 使用绝对路径
            excel_path_abs = os.path.abspath(excel_path)
            output_pdf_abs = os.path.abspath(output_pdf)
            
            # 启动 Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Save and set ActivePrinter for paper size support
            original_printer = None
            try:
                original_printer = excel.ActivePrinter
                # Try "Microsoft Print to PDF" first (supports A3)
                for candidate in ["Microsoft Print to PDF", "Microsoft Print to PDF on Ne01:"]:
                    try:
                        excel.ActivePrinter = candidate
                        logger.info(f"PATH: Excel ActivePrinter set to: {candidate}")
                        break
                    except Exception:
                        continue
                if original_printer is None:
                    logger.warning("Failed to set ActivePrinter, paper size may be limited")
            except Exception as _ap_e:
                logger.warning(f"Failed to save/set ActivePrinter: {_ap_e}")
            
            try:
                # 打开工作簿
                workbook = excel.Workbooks.Open(excel_path_abs)
                
                # 设置打印质量为高质量
                # 遍历所有工作表设置打印区域
                for sheet in workbook.Worksheets:
                    sheet.PageSetup.Zoom = False
                    sheet.PageSetup.FitToPagesWide = 1
                    sheet.PageSetup.FitToPagesTall = False
                    sheet.PageSetup.PrintGridlines = True
                    sheet.PageSetup.CenterHorizontally = True
                    sheet.PageSetup.CenterVertically = False
                
                # Apply user-selected page size
                if page_size:
                    try:
                        w, h = page_size
                        ps = None
                        if abs(w - 595.28) < 5 and abs(h - 841.89) < 5:
                            ps = 9  # xlPaperA4
                        elif abs(w - 841.89) < 5 and abs(h - 1190.55) < 5:
                            ps = 8  # xlPaperA3
                        elif abs(w - 612) < 5 and abs(h - 792) < 5:
                            ps = 1  # xlPaperLetter
                        if ps is not None:
                            for sheet in workbook.Worksheets:
                                sheet.PageSetup.PaperSize = ps
                            logger.info(f"PATH: Excel PaperSize applied, ps={ps}")
                    except Exception as _ps_e:
                        logger.warning(f"Failed to set Excel paper size: {_ps_e}")
                
                # 导出为 PDF
                # XlFixedFormatType.xlTypePDF = 0
                # XlFixedFormatQuality.xlQualityStandard = 0
                workbook.ExportAsFixedFormat(
                    Type=0,
                    Filename=output_pdf_abs,
                    Quality=0,  # xlQualityStandard
                    IncludeDocProperties=True,
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False
                )
                
                # 关闭工作簿
                workbook.Close(False)
                
                if os.path.exists(output_pdf_abs):
                    return True, ""
                else:
                    return False, "Excel 导出 PDF 失败"
            
            finally:
                # Restore original printer
                if original_printer:
                    try:
                        excel.ActivePrinter = original_printer
                        logger.info(f"PATH: Excel ActivePrinter restored to: {original_printer[:60]}...")
                    except Exception:
                        pass
                excel.Quit()
        
        finally:
            pythoncom.CoUninitialize()
    
    except ImportError:
        return False, "缺少 pywin32 库"
    
    except Exception as e:
        error_msg = str(e)
        if "Excel.Application" in error_msg:
            return False, "未检测到 Microsoft Excel"
        return False, f"Excel COM 转换失败: {error_msg}"


def _excel_to_pdf_libreoffice(
    excel_path: str,
    output_dir: str,
    page_size: tuple = None
) -> tuple[bool, str]:
    """
    使用 LibreOffice 命令行转换 Excel 到 PDF
    
    Args:
        excel_path: Excel 文件路径
        output_dir: 输出目录
        page_size: 页面大小 (LibreOffice CLI 不支持)
    
    Returns:
        (是否成功, 错误消息)
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info("PATH: Excel LibreOffice fallback, page_size not supported by CLI")
    try:
        # 查找 LibreOffice
        soffice_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        
        soffice = None
        for path in soffice_paths:
            if os.path.exists(path):
                soffice = path
                break
        
        # 尝试从 PATH 查找
        if not soffice:
            result = subprocess.run(['where', 'soffice'], capture_output=True, text=True,
                                   creationflags=_CREATE_NO_WINDOW)
            if result.returncode == 0:
                soffice = 'soffice'
        
        if not soffice:
            return False, "未找到 Microsoft Excel 或 LibreOffice，请安装其中一个"
        
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 执行转换
        cmd = [
            soffice,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120,
                               creationflags=_CREATE_NO_WINDOW)
        
        if result.returncode == 0:
            # 检查输出文件
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            output_pdf = os.path.join(output_dir, f"{base_name}.pdf")
            
            if os.path.exists(output_pdf):
                return True, ""
            else:
                return False, "转换完成但未找到输出文件"
        
        return False, f"LibreOffice 转换失败: {result.stderr}"
    
    except subprocess.TimeoutExpired:
        return False, "LibreOffice 转换超时（120秒）"
    
    except FileNotFoundError:
        return False, "LibreOffice 命令未找到"
    
    except Exception as e:
        return False, f"转换错误: {e}"


def excel_to_pdf_images(
    excel_path: str,
    output_pdf: str,
    page_size: tuple = None
) -> tuple[bool, str]:
    """
    使用 openpyxl + Pillow 将 Excel 转换为 PDF（备用方案，效果有限）
    
    注意：此方法只能渲染简单的表格，无法完美还原 Excel 格式
    
    Args:
        excel_path: Excel 文件路径
        output_pdf: 输出 PDF 路径
    
    Returns:
        (是否成功, 错误消息)
    """
    try:
        from openpyxl import load_workbook
        from PIL import Image, ImageDraw, ImageFont
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info("PATH: Excel Pillow fallback, page_size not supported (fixed px rendering)")
        
        # 加载 Excel 文件
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        images = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 计算表格尺寸
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 0 or max_col == 0:
                continue
            
            # 创建图片（简单渲染）
            cell_width = 100
            cell_height = 30
            img_width = max_col * cell_width
            img_height = max_row * cell_height
            
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)
            
            # 绘制表格线和内容
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_row, max_col=max_col)):
                y = row_idx * cell_height
                for col_idx, cell in enumerate(row):
                    x = col_idx * cell_width
                    
                    # 绘制单元格边框
                    draw.rectangle(
                        [x, y, x + cell_width, y + cell_height],
                        outline='black'
                    )
                    
                    # 绘制单元格内容
                    if cell.value is not None:
                        text = str(cell.value)[:20]  # 截断长文本
                        draw.text((x + 5, y + 5), text, fill='black')
            
            images.append(img)
        
        wb.close()
        
        if not images:
            return False, "Excel 文件为空"
        
        # 保存为 PDF
        first_img = images[0]
        other_imgs = images[1:] if len(images) > 1 else []
        
        first_img.save(
            output_pdf,
            'PDF',
            save_all=True,
            append_images=other_imgs
        )
        
        return True, ""
    
    except ImportError as e:
        return False, f"缺少依赖库: {e}"
    
    except Exception as e:
        return False, f"转换错误: {e}"
